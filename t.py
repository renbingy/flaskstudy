from openpyxl import load_workbook

workbook_path = 'D:\myflask\wqsl2.xlsx'
from openpyxl import load_workbook
def delete_rows_with_value(workbook_path, value):
    # 加载工作簿
    workbook = load_workbook(workbook_path)
    # 遍历每个sheet页
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # 创建一个空列表，用于存储需要删除的行号
        rows_to_delete = []
        # 遍历每一行
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
            # 判断第一列的值是否为1
            if row[0].value == value:
                # 将需要删除的行号添加到列表中
                rows_to_delete.append(row[0].row)
        # 逆序删除行
        for row_num in reversed(rows_to_delete):
            sheet.delete_rows(row_num)
    # 保存修改后的工作簿
    workbook.save(workbook_path)

    workbook.save(workbook_path)

delete_rows_with_value(workbook_path, 1)